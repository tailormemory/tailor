"""
Estrazione xlsx header-aware — detection conservativa + fallback byte-identico.

Le fixture sono generate al volo con openpyxl (nessun binario nel repo).
Il test critico è `test_unstructured_sheet_byte_identical`: se la detection
sbaglia, l'output di TUTTI i fogli non tabellari cambia silenziosamente.
"""

import datetime
import os
import sys

import pytest

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(BASE_DIR, "scripts", "ingest"))
sys.path.insert(0, os.path.join(BASE_DIR, "scripts", "lib"))

openpyxl = pytest.importorskip("openpyxl")

import ingest_docs  # noqa: E402


def _make_xlsx(tmp_path, name, sheets):
    """sheets: {sheet_name: [[cell, ...], ...]}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return str(path)


def _legacy_reference(path):
    """Reimplementazione letterale dell'estrattore pre-modifica (l'oracolo)."""
    from openpyxl import load_workbook

    sections = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            sections.append({
                "text": f"[Sheet: {sheet_name}]\n" + "\n".join(rows),
                "metadata": {"sheet": sheet_name, "total_sheets": len(wb.sheetnames),
                             "row_count": len(rows)},
            })
    wb.close()
    return sections


# --- detection positiva -----------------------------------------------------

def test_normal_header_detected(tmp_path):
    path = _make_xlsx(tmp_path, "normale.xlsx", {"Costi": [
        ["Voce", "Importo", "Data"],
        ["Pulizie", 45.0, datetime.datetime(2026, 3, 1)],
        ["Lavanderia", 30.5, datetime.datetime(2026, 3, 2)],
        ["Check-in", 20.0, datetime.datetime(2026, 3, 3)],
    ]})
    sections = ingest_docs.extract_excel(path)

    assert len(sections) == 1
    text = sections[0]["text"]
    assert text.startswith("[Sheet: Costi]\nHeader: Voce | Importo | Data")
    assert "Pulizie: Importo=45; Data=2026-03-01" in text
    assert "Lavanderia: Importo=30.5; Data=2026-03-02" in text
    assert sections[0]["metadata"]["prechunked"] is True
    # La riga header non deve comparire anche come riga dati.
    assert "Voce: Importo=Importo" not in text


def test_title_row_before_header(tmp_path):
    """Una riga-titolo (cella singola) può precedere l'header."""
    path = _make_xlsx(tmp_path, "titolo.xlsx", {"Report": [
        ["Riepilogo trimestrale 2026"],
        ["Mese", "Ricavi", "Notti"],
        ["Gennaio", 1200.0, 14.0],
        ["Febbraio", 980.0, 11.0],
        ["Marzo", 1430.0, 17.0],
    ]})
    sections = ingest_docs.extract_excel(path)

    text = sections[0]["text"]
    assert "Header: Mese | Ricavi | Notti" in text
    assert "Gennaio: Ricavi=1200; Notti=14" in text
    # Il titolo sta sopra l'header: v1 accetta di perderlo dal corpo dati.
    assert "Riepilogo trimestrale" not in text


def test_header_repeated_on_every_section(tmp_path):
    """Sezioni pre-chunked: header su OGNI blocco, righe mai spezzate."""
    rows = [["Voce", "Importo", "Nota"]]
    rows += [[f"Voce numero {i}", float(i), "descrizione " + "x" * 60]
             for i in range(200)]
    path = _make_xlsx(tmp_path, "lungo.xlsx", {"Dati": rows})

    sections = ingest_docs.extract_excel(path)
    assert len(sections) > 1
    for sec in sections:
        assert sec["text"].startswith("[Sheet: Dati]\nHeader: Voce | Importo | Nota")
        assert sec["metadata"]["prechunked"] is True
        assert len(sec["text"]) <= ingest_docs.TARGET_CHUNK_CHARS + 200
        # Nessuna riga troncata a metà: ogni riga dati è completa.
        for line in sec["text"].split("\n")[2:]:
            assert line.startswith("Voce numero ")
            assert line.endswith("x" * 10)


# --- fallback ---------------------------------------------------------------

def test_numeric_first_row_falls_back(tmp_path):
    """Prima riga numerica = nessun header -> formato legacy."""
    path = _make_xlsx(tmp_path, "numerico.xlsx", {"Grezzo": [
        [1.0, 2.0, 3.0],
        [4.0, 5.0, 6.0],
        [7.0, 8.0, 9.0],
    ]})
    sections = ingest_docs.extract_excel(path)

    assert sections == _legacy_reference(path)
    assert "Header:" not in sections[0]["text"]
    assert "prechunked" not in sections[0]["metadata"]


def test_unstructured_sheet_byte_identical(tmp_path):
    """Foglio di prosa: nessuna colonna tipizzata -> output byte-identico."""
    path = _make_xlsx(tmp_path, "prosa.xlsx", {"Note": [
        ["Appunti riunione", "presenti Emiliano e Marco"],
        ["Discusso il rinnovo", "da confermare via email"],
        ["Prossimi passi", "preparare il preventivo"],
    ]})
    sections = ingest_docs.extract_excel(path)

    assert sections == _legacy_reference(path)


def test_multi_sheet_mixed_modes(tmp_path):
    """Detection per-foglio: tabellare header-aware, prosa legacy."""
    path = _make_xlsx(tmp_path, "misto.xlsx", {
        "Tabella": [["Voce", "Importo"], ["A", 1.0], ["B", 2.0], ["C", 3.0]],
        "Prosa": [["Nota libera qui", "altro testo"], ["Seconda riga", "ancora testo"]],
    })
    sections = ingest_docs.extract_excel(path)
    by_sheet = {s["metadata"]["sheet"]: s for s in sections}

    assert "Header: Voce | Importo" in by_sheet["Tabella"]["text"]
    legacy = {s["metadata"]["sheet"]: s for s in _legacy_reference(path)}
    assert by_sheet["Prosa"] == legacy["Prosa"]


def test_single_data_row_falls_back(tmp_path):
    """Evidenza insufficiente (1 sola riga dati) -> conservativo, legacy."""
    path = _make_xlsx(tmp_path, "unariga.xlsx", {"S": [
        ["Voce", "Importo"],
        ["Pulizie", 45.0],
    ]})
    assert ingest_docs.extract_excel(path) == _legacy_reference(path)


def test_data_wider_than_header_falls_back(tmp_path):
    """Dati che debordano oltre l'ultima colonna intestata -> non è un header."""
    path = _make_xlsx(tmp_path, "debordo.xlsx", {"S": [
        ["Voce", "Importo", None],
        ["Pulizie", 45.0, 10.0],
        ["Lavanderia", 30.0, 12.0],
    ]})
    assert ingest_docs.extract_excel(path) == _legacy_reference(path)


def test_multi_block_sheet_falls_back(tmp_path):
    """Foglio con più tabelle, ognuna col proprio header (tipico P&L:
    GROSS INCOME / EXPENSES / ASSETS). Applicare il primo header a tutto il
    foglio produrrebbe righe assurde tipo "EXPENSES: Gennaio=Gennaio"."""
    path = _make_xlsx(tmp_path, "multiblocco.xlsx", {"2023": [
        ["GROSS INCOME", "Gennaio", "Febbraio", "Marzo"],
        ["Stipendio", 6167.0, 6167.0, 6167.0],
        ["Dividendi", 0.0, 0.0, 32900.0],
        ["TOTAL", 6167.0, 6167.0, 39067.0],
        ["EXPENSES", "Gennaio", "Febbraio", "Marzo"],
        ["Mantenimento", 5000.0, 5000.0, 5000.0],
        ["Viaggi", 7200.0, 1780.0, 7270.0],
    ]})
    sections = ingest_docs.extract_excel(path)

    assert sections == _legacy_reference(path)
    joined = sections[0]["text"]
    assert "Gennaio=Gennaio" not in joined


def test_partial_header_repeat_does_not_trigger_guard(tmp_path):
    """Una singola cella che ripete un'etichetta di header non è un blocco
    nuovo: la guardia richiede >=2 ripetizioni sulla stessa riga."""
    path = _make_xlsx(tmp_path, "ripetizione.xlsx", {"S": [
        ["Categoria", "Importo", "Mese"],
        ["Importo", 100.0, "Gennaio"],   # "Importo" ripetuto solo in col 0
        ["Pulizie", 45.0, "Febbraio"],
        ["Viaggi", 300.0, "Marzo"],
    ]})
    sections = ingest_docs.extract_excel(path)

    assert "Header: Categoria | Importo | Mese" in sections[0]["text"]
    assert sections[0]["metadata"]["prechunked"] is True


# --- limite noto: header multi-riga ----------------------------------------

def test_multirow_header_documented_behavior(tmp_path):
    """v1 NON unisce header multi-riga: documenta il comportamento effettivo.

    Riga 0 ("Costi | | Ricavi | ") e riga 1 ("Voce | Importo | Fonte | Totale")
    formano un header a due livelli. La riga 0 viene eletta candidata (2 celle
    testuali), ma il suo span si ferma a "Ricavi" (col 2) mentre i dati
    occupano anche la col 3 → debordo → fallback legacy. Nessuna corruzione,
    solo il formato vecchio.
    """
    path = _make_xlsx(tmp_path, "multirow.xlsx", {"S": [
        ["Costi", None, "Ricavi", None],
        ["Voce", "Importo", "Fonte", "Totale"],
        ["Pulizie", 45.0, "Airbnb", 120.0],
        ["Lavanderia", 30.0, "Booking", 90.0],
    ]})
    sections = ingest_docs.extract_excel(path)

    # Comportamento atteso in v1: fallback legacy, nessuna corruzione.
    assert sections == _legacy_reference(path)


# --- helper unitari ---------------------------------------------------------

@pytest.mark.parametrize("value,expected", [
    ("Voce", True),
    ("Importo netto", True),
    ("", False),
    (None, False),
    (42, False),
    (3.14, False),
    (True, False),
    ("42", False),
    ("1.234,56", False),
    ("-3%", False),
    ("€ 1.200", False),
    ("2026-03-01", False),
    ("31/12/24", False),
    ("---", False),
    (datetime.datetime(2026, 3, 1), False),
    (datetime.date(2026, 3, 1), False),
])
def test_is_texty(value, expected):
    assert ingest_docs._xl_is_texty(value) is expected


@pytest.mark.parametrize("value,expected", [
    (3.0, "3"),
    (30.5, "30.5"),
    (None, ""),
    (True, "true"),
    (datetime.datetime(2026, 3, 1), "2026-03-01"),
    (datetime.datetime(2026, 3, 1, 14, 30), "2026-03-01 14:30:00"),
    ("  spazi  ", "spazi"),
])
def test_fmt_value(value, expected):
    assert ingest_docs._xl_fmt(value) == expected


# --- integrazione col chunker ----------------------------------------------

def test_prechunked_sections_bypass_splitter_and_overlap(tmp_path):
    rows = [["Voce", "Importo", "Nota"]]
    rows += [[f"Riga {i}", float(i), "y" * 80] for i in range(100)]
    path = _make_xlsx(tmp_path, "chunk.xlsx", {"Dati": rows})
    sections = ingest_docs.extract_excel(path)

    file_info = {
        "filepath": str(path), "filename": "chunk.xlsx", "rel_path": "chunk.xlsx",
        "hash": "a" * 64, "modified_date": "2026-07-19", "modified_ts": 0,
        "extension": ".xlsx", "category": "",
    }
    chunks = ingest_docs.chunk_document(sections, file_info)

    assert len(chunks) == len(sections)
    for chunk, sec in zip(chunks, sections):
        assert chunk["text"] == sec["text"]              # nessuno split
        assert "[...contesto precedente:]" not in chunk["text"]  # nessun overlap
        assert "sec_prechunked" not in chunk["metadata"]  # non finisce in Chroma
        assert chunk["metadata"]["sec_sheet"] == "Dati"
